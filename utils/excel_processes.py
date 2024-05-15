from openpyxl import load_workbook
from io import BytesIO
from aiogram import Bot
from aiogram.types import Message
from typing import BinaryIO


async def read_excel(bot: Bot, message: Message):
    xl_bytesio: BinaryIO = await bot.download(message.bot, file=message.document)
    xl_bytes: bytes = xl_bytesio.getvalue()
    wb = load_workbook(filename=BytesIO(xl_bytes))
    ws = wb[wb.sheetnames[0]]
    result = []
    columns = [c.value for row in ws.iter_rows(min_row=1, max_col=ws.max_column, max_row=1) for c in row]
    for row in ws.iter_rows(min_row=2, max_col=ws.max_column, max_row=ws.max_row):
        result.append({columns[k]: str(v.value) for k, v in enumerate(row)})
    return result
